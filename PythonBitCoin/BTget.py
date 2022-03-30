# -*- coding: euc-kr -*-

from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
from win32con import TRUE
import pyupbit
import numpy as np

path = 'data.xlsx'

wb = Workbook()
sheet1 = wb.active
sheet2 = wb.active


def write_exc(script, logtype):
    wb = Workbook()
    sheet1 = wb.active
    sheet2 = wb.active

    sheet1.title = "TransAction"
    sheet2.title = "Profit"

    if logtype == "BuyCoin" or logtype == "SellCoin":
        sheet1.append([str(datetime.datetime.now()), ticker, volume, script, logtype])
    elif logtype == "ProfitReport":
        sheet2.append([str(datetime.datetime.now()), script, logtype])

    wb.save(path)

